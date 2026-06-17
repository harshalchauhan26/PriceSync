#!/usr/bin/env python3
"""
PriceSync — one app, SQLite-backed.

Pages:
  /pipeline       run the price checker (DB in, DB out — no Excel locks)
  /review         approve mismatches, re-run errors, push prices to Shopify
  /integrations   add a Shopify token per brand and it just works

All state lives in pricesync.db (see store.py). The Excel sheet is only ever
read once, on import. Run:  python saas.py  ->  http://127.0.0.1:8080
"""

import io
import os
import random
import smtplib
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from email.message import EmailMessage
from urllib.parse import urlsplit

import pandas as pd
import requests
from flask import Flask, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename

import fx
import price_tracker as pt
import store


def _match_tol(base, cur):
    """Foreign-currency conversions carry FX noise, so allow a small relative
    band for non-INR; INR stays strict (legacy behaviour)."""
    if cur in ("INR", "UNKNOWN", None, ""):
        return pt.MATCH_TOLERANCE
    return max(pt.MATCH_TOLERANCE, 0.005 * abs(base or 0))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = int(
    float(os.environ.get("MAX_UPLOAD_MB", "64")) * 1024 * 1024)
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "uploads")

LOCK = threading.Lock()
LOG = []
LOG_MAX = 5000                 # keep memory bounded; older entries are dropped
LOGMETA = {"offset": 0}        # absolute index of LOG[0] (entries dropped so far)
_thread_local = threading.local()
ACTIVE_COOLDOWN = pt.COOLDOWN_RANGE

CONFIG = {
    "concurrency": 5, "timeout_ms": 12000, "batch_size": 500, "rest_between": 60,
    "simulation": False, "retry_errors": False, "fresh_start": False,
    "safe_retry": True, "safe_concurrency": 1, "safe_cooldown_min": 4.0,
    "safe_cooldown_max": 8.0, "safe_rest_between": 30, "safe_batch_size": 25,
    "vendors": [],   # empty = whole catalog; else only these vendor domains
}
STATE = {
    "running": False, "abort": False, "phase": "idle", "total_rows": 0,
    "pre_done": 0, "completed": 0, "matched": 0, "mismatch": 0, "errors": 0,
    "retry_total": 0, "retry_completed": 0, "retry_recovered": 0,
    "started_at": None, "message": "Idle. Import a sheet, then Start.",
}
RERUN = {"running": False, "total": 0, "done": 0, "recovered": 0,
         "still_error": 0, "message": "Idle."}
RERUN_LOCK = threading.Lock()
PUSH = {"running": False, "total": 0, "done": 0, "ok": 0, "failed": 0, "message": "Idle."}
PUSH_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Engine glue (DB-backed)
# ---------------------------------------------------------------------------

def _get_fetcher():
    f = getattr(_thread_local, "fetcher", None)
    if f is None:
        f = pt.Fetcher(timeout=CONFIG["timeout_ms"] / 1000.0,
                       cooldown_range=ACTIVE_COOLDOWN, quiet=True)
        _thread_local.fetcher = f
    return f


def _log(row_no, platform, brand, currency, price, status, msg="", url=""):
    with LOCK:
        LOG.append({
            "row": row_no, "platform": (platform or "?").lower(), "domain": brand,
            "url": url, "currency": currency or "-",
            "price": f"{price:.2f}" if isinstance(price, (int, float)) else "-",
            "status": status, "msg": msg})
        if len(LOG) > LOG_MAX:               # bound memory; drop oldest in bulk
            drop = len(LOG) - LOG_MAX
            del LOG[:drop]
            LOGMETA["offset"] += drop


def _process(prod):
    """Fetch + verify one product (a dict read from the sheet).
    Returns (prod, status, live, cur, state)."""
    url = (prod["url"] or "").strip()
    platform = (prod["platform"] or "").strip()
    regex = prod["custom_regex"] or None
    base = prod["base_price"]
    brand = prod["brand"]
    tag = prod.get("key") or url

    if CONFIG["simulation"]:
        time.sleep(random.uniform(0.03, 0.12))
        roll = random.random()
        if roll < 0.05 or base is None:
            _log(tag, platform, brand, None, None, "Fetch Error",
                 "simulated failure", url=url)
            return prod, "Fetch Error", None, None, "error"
        live = base if roll > 0.12 else round(base * random.choice([0.9, 1.1]), 2)
        cur = "INR" if brand.endswith(".in") else "USD"
        if abs(live - base) <= pt.MATCH_TOLERANCE:
            _log(tag, platform, brand, cur, live, "Price Matched", url=url)
            return prod, f"Price Matched ({cur})", live, cur, "matched"
        _log(tag, platform, brand, cur, live, "Price Mismatch!",
             f"delta {live - base:+.2f}", url=url)
        return prod, f"Price Mismatch! ({cur})", live, cur, "mismatch"

    try:
        live, currency = pt.extract_row(_get_fetcher(), url, platform, regex)
        if live is None:
            raise ValueError("price not found")
    except Exception as exc:
        _log(tag, platform, brand, None, None, "Fetch Error", f"{exc}", url=url)
        return prod, "Fetch Error", None, None, "error"

    cur = currency or "UNKNOWN"
    if base is None:
        _log(tag, platform, brand, cur, live, "Fetch Error",
             "baseline price unreadable", url=url)
        return prod, "Fetch Error", live, cur, "error"
    live_inr = fx.to_inr(live, cur)          # USD/CAD -> INR at current rate
    delta = live_inr - base                  # compare in INR
    disp = cur if cur in ("INR", "UNKNOWN") else f"{cur}->INR"
    note = "" if cur in ("INR", "UNKNOWN") else f"{cur} {live:g} @{fx.rate_of(cur):.2f}"
    if abs(delta) <= _match_tol(base, cur):
        _log(tag, platform, brand, disp, live_inr, "Price Matched", note, url=url)
        return prod, f"Price Matched ({cur})", live, cur, "matched"
    _log(tag, platform, brand, disp, live_inr, "Price Mismatch!",
         (note + f" · Δ{delta:+.0f}").strip(" ·"), url=url)
    return prod, f"Price Mismatch! ({cur})", live, cur, "mismatch"


_RESULT_UPSERT = """
INSERT INTO products
  (key, mbo_url, url, platform, custom_regex, brand, base_price,
   live_price, currency, status, state, delta, updated_at)
VALUES (%(key)s,%(mbo_url)s,%(url)s,%(platform)s,%(custom_regex)s,%(brand)s,
        %(base_price)s,%(live_price)s,%(currency)s,%(status)s,%(state)s,
        %(delta)s,%(updated_at)s)
ON CONFLICT(key) DO UPDATE SET
  live_price=excluded.live_price, currency=excluded.currency,
  status=excluded.status, state=excluded.state, delta=excluded.delta,
  updated_at=excluded.updated_at
"""


def _save_result(con, prod, status, live, cur, state, run_id):
    """Write the latest result to `products` (upsert by key) and append a row
    to `price_history` so the Alerts page can see movement over time."""
    base = prod["base_price"]
    # delta is in INR: convert the (raw) live price by its currency first
    delta = (fx.to_inr(live, cur) - base) if (live is not None and base is not None) else None
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    con.execute(_RESULT_UPSERT, {
        "key": prod["key"], "mbo_url": prod.get("mbo_url", ""), "url": prod["url"],
        "platform": prod["platform"], "custom_regex": prod["custom_regex"],
        "brand": prod["brand"], "base_price": base, "live_price": live,
        "currency": cur, "status": status, "state": state, "delta": delta,
        "updated_at": now})
    con.execute(
        "INSERT INTO price_history(key,url,brand,base_price,live_price,delta,"
        "state,status,run_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (prod["key"], prod["url"], prod["brand"], base, live, delta,
         state, status, run_id))
    return delta


def _abortable_rest(seconds, label):
    with LOCK:
        STATE["message"] = f"{label} - resting {max(0, seconds):.0f}s..."
    t0 = time.time()
    while time.time() - t0 < max(0.0, float(seconds)) and not STATE["abort"]:
        time.sleep(0.5)


def _run_pass(con, rows, workers, batch_size, rest_between, cooldown, phase,
              label, on_done, run_id):
    """rows is a list of product dicts read from the sheet."""
    global ACTIVE_COOLDOWN
    ACTIVE_COOLDOWN = cooldown
    workers = max(1, int(workers))
    batch_size = max(1, int(batch_size))
    with LOCK:
        STATE["phase"] = phase
    for bstart in range(0, len(rows), batch_size):
        if STATE["abort"]:
            return False
        batch = rows[bstart:bstart + batch_size]
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(_process, p) for p in batch]
            for fut in as_completed(futs):
                if STATE["abort"]:
                    for f2 in futs:
                        f2.cancel()
                    break
                prod, status, live, cur, st = fut.result()
                _save_result(con, prod, status, live, cur, st, run_id)
                con.commit()  # commit per row so Review/counts update live
                on_done(prod, st)
        if not STATE["abort"] and bstart + batch_size < len(rows):
            _abortable_rest(rest_between, label)
            with LOCK:
                STATE["message"] = label
    return not STATE["abort"]


def _pipeline():
    global ACTIVE_COOLDOWN
    con = store.connect()
    run_id = time.strftime("%Y%m%d-%H%M%S")
    try:
        mode = "fresh" if CONFIG["fresh_start"] else "update"
        vendors = CONFIG.get("vendors") or None      # vendor-scoped run
        rows = store.work_rows(mode, vendors)
        total = store.count_products(vendors)
        scope = f"{len(vendors)} vendor(s)" if vendors else "all vendors"
        with LOCK:
            STATE.update(total_rows=total, pre_done=max(0, total - len(rows)),
                         message=f"Main pass - {len(rows)} product(s) · {scope}")

        def main_done(prod, st):
            with LOCK:
                STATE["completed"] += 1
                STATE["matched" if st == "matched" else
                      "mismatch" if st == "mismatch" else "errors"] += 1

        finished = _run_pass(
            con, rows, min(25, max(1, int(CONFIG["concurrency"]))),
            CONFIG["batch_size"], CONFIG["rest_between"], pt.COOLDOWN_RANGE,
            "main", "Main pass", main_done, run_id)

        if finished and CONFIG["safe_retry"]:
            # which of the worked links ended this run in error (with a baseline)
            erkeys = {r["key"] for r in con.execute(
                "SELECT key FROM products WHERE state='error' "
                "AND base_price IS NOT NULL").fetchall()}
            err = [p for p in rows if p["key"] in erkeys]
            with LOCK:
                STATE.update(retry_total=len(err), retry_completed=0, retry_recovered=0)
            if err:
                with LOCK:
                    STATE["message"] = f"Safe-Retry Window - {len(err)} errors, gently"

                def retry_done(prod, st):
                    with LOCK:
                        STATE["retry_completed"] += 1
                        if st != "error":
                            STATE["retry_recovered"] += 1
                            STATE["errors"] = max(0, STATE["errors"] - 1)
                            if st == "matched":
                                STATE["matched"] += 1
                            elif st == "mismatch":
                                STATE["mismatch"] += 1

                _run_pass(con, err, CONFIG["safe_concurrency"],
                          CONFIG["safe_batch_size"], CONFIG["safe_rest_between"],
                          (float(CONFIG["safe_cooldown_min"]),
                           float(CONFIG["safe_cooldown_max"])),
                          "safe_retry", "Safe-Retry Window", retry_done, run_id)
        ACTIVE_COOLDOWN = pt.COOLDOWN_RANGE
        final = "Aborted" if STATE["abort"] else "Completed"
        rec = STATE["retry_recovered"]
        with LOCK:
            STATE["phase"] = "done"
            STATE["message"] = (f"{final}" +
                                (f" ({rec} recovered in Safe-Retry Window)" if rec else "") +
                                ". Results saved to database.")
    finally:
        con.close()
        with LOCK:
            STATE["running"] = False


# ---------------------------------------------------------------------------
# Shopify
# ---------------------------------------------------------------------------

def _integration(brand):
    con = store.connect()
    r = con.execute("SELECT * FROM integrations WHERE brand=%s", (brand,)).fetchone()
    con.close()
    return dict(r) if r else None


def _product_handle(url):
    return urlsplit(url).path.rstrip("/").rsplit("/", 1)[-1].split("?")[0]


def push_price_to_shopify(brand, url, price):
    cfg = _integration(brand)
    if not cfg or not cfg.get("shop_domain") or not cfg.get("access_token"):
        return {"ok": False, "status": f"no Shopify token for '{brand}' (add it in Integrations)"}
    if price is None:
        return {"ok": False, "status": "approve a final price first"}
    handle = _product_handle(url)
    if cfg.get("dry_run"):
        return {"ok": True, "status": f"DRY RUN: would set '{handle}' -> {price}"}
    ver = cfg.get("api_version") or "2024-10"
    base = f"https://{cfg['shop_domain']}/admin/api/{ver}"
    headers = {"X-Shopify-Access-Token": cfg["access_token"], "Content-Type": "application/json"}
    try:
        r = requests.get(f"{base}/products.json",
                         params={"handle": handle, "fields": "id,variants"},
                         headers=headers, timeout=20)
        r.raise_for_status()
        products = r.json().get("products", [])
        if not products:
            return {"ok": False, "status": f"handle '{handle}' not found in store"}
        variants = products[0].get("variants", [])
        updated = 0
        for v in variants:
            pr = requests.put(f"{base}/variants/{v['id']}.json",
                              json={"variant": {"id": v["id"], "price": f"{price}"}},
                              headers=headers, timeout=20)
            if pr.status_code in (200, 201):
                updated += 1
        ok = updated > 0
        return {"ok": ok, "status": f"{'updated' if ok else 'FAILED'} "
                f"{updated}/{len(variants)} variant(s) -> {price}"}
    except requests.RequestException as exc:
        return {"ok": False, "status": f"Shopify API error: {exc}"}


def verify_shopify(brand):
    cfg = _integration(brand)
    if not cfg or not cfg.get("shop_domain") or not cfg.get("access_token"):
        return {"ok": False, "status": f"no token saved for '{brand}'"}
    ver = cfg.get("api_version") or "2024-10"
    try:
        r = requests.get(f"https://{cfg['shop_domain']}/admin/api/{ver}/shop.json",
                         headers={"X-Shopify-Access-Token": cfg["access_token"]}, timeout=15)
    except requests.RequestException as exc:
        return {"ok": False, "status": f"connection error: {exc}"}
    if r.status_code == 200:
        shop = r.json().get("shop", {})
        mode = " · DRY-RUN" if cfg.get("dry_run") else " · LIVE"
        return {"ok": True, "status": f"connected to {shop.get('name')} ({shop.get('myshopify_domain')}){mode}"}
    if r.status_code == 401:
        return {"ok": False, "status": "401 - invalid/expired token"}
    if r.status_code == 404:
        return {"ok": False, "status": "404 - shop_domain not found"}
    return {"ok": False, "status": f"HTTP {r.status_code}"}


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/")
def home():
    # Served raw (not via Jinja) — the React/JSX file is full of {{ }} braces.
    return send_file(os.path.join(app.root_path, "templates", "dashboard.html"))


@app.route("/classic")
def home_classic():
    return render_template("pipeline.html", active="pipeline")


@app.route("/pipeline")
def page_pipeline():
    return render_template("pipeline.html", active="pipeline")


@app.route("/review")
def page_review():
    return render_template("review.html", active="review")


@app.route("/integrations")
def page_integrations():
    return render_template("integrations.html", active="integrations")


@app.route("/alerts")
def page_alerts():
    return render_template("alerts.html", active="alerts")


@app.route("/healthz")
def healthz():
    with LOCK:
        return jsonify(ok=True, running=STATE["running"], phase=STATE["phase"])


@app.route("/api/meta")
def api_meta():
    return jsonify(counts=store.counts(),
                   alerts=store.alert_count(5),
                   last_import=store.get_meta("last_import"),
                   last_import_rows=store.get_meta("last_import_rows"),
                   last_import_file=store.get_meta("last_import_file"))


# ---- Pipeline API ----

def _save_upload(f):
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    path = os.path.join(UPLOAD_DIR, secure_filename(f.filename))
    f.save(path)
    return path


@app.route("/api/import/preview", methods=["POST"])
def api_import_preview():
    """Inspect an uploaded sheet (no DB write): return distinct designer
    domains + counts so the user can filter before importing."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(ok=False, error="no file"), 400
    path = _save_upload(f)
    try:
        prev = store.preview_sheet(path)
    except Exception as exc:
        return jsonify(ok=False, error=str(exc)), 400
    # stash the path so /api/import can reuse it without re-uploading
    return jsonify(ok=True, path=os.path.basename(path), **prev)


@app.route("/api/import", methods=["POST"])
def api_import():
    """Import with an optional Designer-URL filter (contains text + domain
    whitelist). Accepts either a fresh file upload or a previously previewed
    filename (form field `path`)."""
    contains = (request.form.get("contains") or request.args.get("contains") or "").strip()
    domains = request.form.getlist("domains") or request.form.getlist("domains[]")
    domains = [d for d in domains if d and d != "(none)"]

    f = request.files.get("file")
    if f and f.filename:
        path = _save_upload(f)
    else:
        name = secure_filename(request.form.get("path", ""))
        path = os.path.join(UPLOAD_DIR, name) if name else None
        if not path or not os.path.exists(path):
            return jsonify(ok=False, error="no file"), 400

    replace = (request.form.get("mode", request.args.get("mode", "replace")) != "add")
    try:
        res = store.import_sheet(path, replace=replace,
                                 contains=contains or None,
                                 domains=domains or None)
    except Exception as exc:
        return jsonify(ok=False, error=str(exc)), 400
    return jsonify(ok=True, **res, counts=store.counts(),
                   contains=contains, domains=domains)


@app.route("/api/clear_db", methods=["POST"])
def api_clear_db():
    if STATE["running"]:
        return jsonify(ok=False, error="stop the running pipeline first"), 409
    con = store.connect()
    con.execute("DELETE FROM products")
    con.commit()
    con.close()
    return jsonify(ok=True, counts=store.counts())


@app.route("/api/pipe/config", methods=["POST"])
def api_config():
    data = request.get_json(force=True) or {}
    with LOCK:
        for k in CONFIG:
            if k in data:
                CONFIG[k] = data[k]
    return jsonify(ok=True, config=CONFIG)


@app.route("/api/pipe/start", methods=["POST"])
def api_start():
    with LOCK:
        if STATE["running"]:
            return jsonify(error="already running"), 409
        if store.counts()["total"] == 0:
            return jsonify(error="no products in database — import a sheet first"), 400
        STATE.update(running=True, abort=False, phase="main", completed=0,
                     matched=0, mismatch=0, errors=0, retry_total=0,
                     retry_completed=0, retry_recovered=0,
                     started_at=time.time(), message="Starting...")
    threading.Thread(target=_pipeline, daemon=True).start()
    return jsonify(ok=True)


@app.route("/api/pipe/abort", methods=["POST"])
def api_abort():
    STATE["abort"] = True
    return jsonify(ok=True)


@app.route("/api/pipe/clear_log", methods=["POST"])
def api_clear_log():
    with LOCK:
        LOG.clear()
        LOGMETA["offset"] = 0
    return jsonify(ok=True)


@app.route("/api/pipe/status")
def api_status():
    cursor = int(request.args.get("cursor", 0))
    with LOCK:
        offset = LOGMETA["offset"]
        start = max(0, cursor - offset)      # cursor is an absolute index
        entries = LOG[start:]
        total = offset + len(LOG)
        snap = dict(STATE)
        cfg = dict(CONFIG)
    elapsed = int(time.time() - snap["started_at"]) if snap["started_at"] else 0
    return jsonify(running=snap["running"], phase=snap["phase"],
                   total_rows=snap["total_rows"], pre_done=snap["pre_done"],
                   completed=snap["completed"],
                   current_row=snap["pre_done"] + snap["completed"],
                   matched=snap["matched"], mismatch=snap["mismatch"],
                   errors=snap["errors"], retry_total=snap["retry_total"],
                   retry_completed=snap["retry_completed"],
                   retry_recovered=snap["retry_recovered"], elapsed=elapsed,
                   message=snap["message"], config=cfg,
                   cursor=total, entries=entries, log_total=total)


_FILL_MISMATCH = "FFF2CC"   # soft yellow
_FILL_ERROR = "F8CBAD"      # soft red
_FILL_HEADER = "1F2A40"


@app.route("/api/export")
def api_export():
    kind = request.args.get("kind", "all")
    fmt_ = request.args.get("fmt", "xlsx")
    con = store.connect()
    where = {"all": "1=1", "mismatch": "state='mismatch'", "error": "state='error'",
             "approved": "decision='approved'"}.get(kind, "1=1")
    cols = ["id", "brand", "platform", "url", "base_price", "live_price",
            "currency", "status", "state", "delta", "decision", "markup_pct",
            "custom_price", "final_price", "shopify_status", "shopify_at"]
    # Build the DataFrame from dict rows directly; pandas' read_sql does not
    # support a raw psycopg3 (dict_row) connection.
    rows = con.execute(
        f"SELECT {', '.join(cols)} FROM products WHERE {where} ORDER BY id"
        ).fetchall()
    con.close()
    df = pd.DataFrame([dict(r) for r in rows], columns=cols)
    bio = io.BytesIO()
    if fmt_ == "csv":
        bio.write(df.to_csv(index=False).encode("utf-8"))
        mime, name = "text/csv", f"pricesync_{kind}.csv"
    else:
        _write_colored_xlsx(bio, df)
        mime, name = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      f"pricesync_{kind}.xlsx")
    bio.seek(0)
    return send_file(bio, mimetype=mime, as_attachment=True, download_name=name)


def _write_colored_xlsx(bio, df):
    """Write df to xlsx and tint each row: yellow for a price MISMATCH,
    red for a fetch ERROR. Color is driven by the `state` column."""
    from openpyxl.styles import PatternFill, Font
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="export")
        ws = xw.sheets["export"]
        try:
            state_col = list(df.columns).index("state") + 1  # 1-based
        except ValueError:
            state_col = None
        head = PatternFill("solid", fgColor=_FILL_HEADER)
        hfont = Font(color="FFFFFF", bold=True)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = head
            ws.cell(row=1, column=c).font = hfont
        yellow = PatternFill("solid", fgColor=_FILL_MISMATCH)
        red = PatternFill("solid", fgColor=_FILL_ERROR)
        if state_col:
            for i in range(len(df)):
                st = df.iat[i, state_col - 1]
                fill = yellow if st == "mismatch" else red if st == "error" else None
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=i + 2, column=c).fill = fill
        # legend on a second sheet
        leg = xw.book.create_sheet("legend")
        leg["A1"] = "Yellow = Price Mismatch"
        leg["A1"].fill = yellow
        leg["A2"] = "Red = Fetch Error"
        leg["A2"].fill = red


# ---- Email the mismatch report ----

def _mismatch_xlsx():
    cols = ["id", "brand", "platform", "url", "base_price", "live_price",
            "currency", "status", "state", "delta", "decision", "final_price"]
    con = store.connect()
    rows = con.execute(
        f"SELECT {', '.join(cols)} FROM products WHERE state='mismatch' "
        "ORDER BY ABS(COALESCE(delta,0)) DESC").fetchall()
    con.close()
    df = pd.DataFrame([dict(r) for r in rows], columns=cols)
    bio = io.BytesIO()
    _write_colored_xlsx(bio, df)
    bio.seek(0)
    return bio.read(), len(df)


def send_mismatch_report(to=None):
    host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER", "").strip()
    pwd = os.environ.get("SMTP_PASS", "").strip()
    frm = (os.environ.get("SMTP_FROM", "") or user).strip()
    to = (to or os.environ.get("ALERT_TO", "")).strip()
    if not user or not pwd:
        return {"ok": False, "error": "email not configured — set SMTP_USER and "
                "SMTP_PASS (Gmail App Password) in .env"}
    if not to:
        return {"ok": False, "error": "no recipient — set ALERT_TO in .env or pass 'to'"}
    data, n = _mismatch_xlsx()
    today = time.strftime("%Y-%m-%d")
    msg = EmailMessage()
    msg["Subject"] = f"PriceSync Alert — {n} price mismatches ({today})"
    msg["From"] = frm
    msg["To"] = to
    msg.set_content(
        f"PriceSync detected {n} price mismatch(es) awaiting review.\n\n"
        "The attached sheet lists every mismatch (yellow rows). These are "
        "PENDING APPROVAL — no price change has been pushed to any store.\n"
        "Review and approve each change in the console before anything goes live.\n\n"
        "— PriceSync (beta)")
    msg.add_attachment(
        data, maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"price_mismatches_{today}.xlsx")
    with smtplib.SMTP(host, port, timeout=30) as s:
        s.ehlo()
        s.starttls()
        s.login(user, pwd)
        s.send_message(msg)
    return {"ok": True, "count": n, "to": to}


@app.route("/api/alerts/email_mismatch", methods=["POST"])
def api_email_mismatch():
    to = ((request.get_json(silent=True) or {}).get("to") or "").strip() or None
    try:
        res = send_mismatch_report(to)
    except Exception as exc:
        return jsonify(ok=False, error=f"send failed: {exc}"), 500
    return jsonify(**res), (200 if res.get("ok") else 400)


# ---- Review API ----

@app.route("/api/review/items")
def api_items():
    kind = request.args.get("kind", "mismatch")
    brand = request.args.get("brand", "").strip()
    brands = [b for b in request.args.get("brands", "").split(",") if b.strip()]
    state = {"mismatch": "mismatch", "error": "error", "resolved": "matched"}.get(kind, "mismatch")
    where, params = "state=%s", [state]
    if brands:  # Excel-style multi-vendor filter
        where += " AND brand IN (" + ",".join(["%s"] * len(brands)) + ")"
        params += brands
    elif brand:
        where += " AND brand=%s"
        params.append(brand)
    con = store.connect()
    rows = con.execute(
        f"SELECT * FROM products WHERE {where} ORDER BY decision='pending' DESC, "
        f"ABS(COALESCE(delta,0)) DESC LIMIT 1000", params).fetchall()
    con.close()
    return jsonify(items=[dict(r) for r in rows], counts=store.counts())


@app.route("/api/review/brands")
def api_brands():
    kind = request.args.get("kind", "").strip()
    state = {"mismatch": "mismatch", "error": "error", "resolved": "matched"}.get(kind)
    con = store.connect()
    if state:
        rows = con.execute("SELECT brand, COUNT(*) c FROM products WHERE state=%s "
                           "AND brand<>'' GROUP BY brand ORDER BY c DESC", (state,)).fetchall()
    else:
        rows = con.execute("SELECT brand, COUNT(*) c FROM products WHERE brand<>'' "
                           "GROUP BY brand ORDER BY c DESC").fetchall()
    configured = {r["brand"] for r in con.execute(
        "SELECT brand FROM integrations WHERE access_token IS NOT NULL AND access_token<>''").fetchall()}
    con.close()
    return jsonify(brands=[{"brand": r["brand"], "count": r["c"],
                            "shopify": r["brand"] in configured} for r in rows])


@app.route("/api/vendors")
def api_vendors():
    """Every distinct vendor (derived from the product links) with a count —
    powers the Excel-style filter. Optional ?kind= limits to a feed."""
    kind = request.args.get("kind", "").strip()
    state = {"mismatch": "mismatch", "error": "error", "resolved": "matched"}.get(kind)
    con = store.connect()
    if state:
        rows = con.execute("SELECT brand, COUNT(*) c FROM products WHERE brand<>'' "
                           "AND state=%s GROUP BY brand ORDER BY brand", (state,)).fetchall()
    else:
        rows = con.execute("SELECT brand, COUNT(*) c FROM products WHERE brand<>'' "
                           "GROUP BY brand ORDER BY brand").fetchall()
    con.close()
    return jsonify(vendors=[{"vendor": r["brand"], "count": r["c"]} for r in rows])


@app.route("/api/review/decide", methods=["POST"])
def api_decide():
    d = request.get_json(force=True) or {}
    pid = d.get("row")
    if pid in (None, ""):
        return jsonify(ok=False, error="missing row"), 400
    con = store.connect()
    it = con.execute("SELECT base_price, live_price, currency FROM products WHERE id=%s",
                     (int(pid),)).fetchone()
    if it is None:
        con.close()
        return jsonify(ok=False, error="unknown row"), 404
    decision = d.get("decision", "approved")
    ref = d.get("ref", "live")
    markup = d.get("markup_pct")
    custom = d.get("custom_price")
    convert = d.get("convert", True)
    markup = None if markup in ("", None) else float(markup)
    custom = None if custom in ("", None) else float(custom)
    final = None
    if decision == "approved":
        final = store_compute_final(it["base_price"], it["live_price"],
                                    it["currency"], ref, markup, custom, convert)
    con.execute("UPDATE products SET decision=%s,markup_pct=%s,custom_price=%s,ref=%s,"
                "final_price=%s,note=%s,decided_at=%s WHERE id=%s",
                (decision, markup, custom, ref, final, d.get("note", ""),
                 time.strftime("%Y-%m-%d %H:%M:%S"), int(pid)))
    con.commit()
    con.close()
    return jsonify(ok=True, final_price=final)


def store_compute_final(base, live, currency, ref, markup_pct, custom_price, convert=True):
    """Final price in INR. custom overrides; else apply markup on the reference.
    When convert=True the live reference is converted from its currency to INR."""
    if custom_price is not None and float(custom_price) > 0:
        return round(float(custom_price), 2)
    if ref == "base":
        reference = base
    else:
        reference = fx.to_inr(live, currency) if convert else live
    if reference is None:
        reference = base
    if reference is None:
        return None
    return round(reference * (1 + float(markup_pct or 0) / 100.0), 2)


@app.route("/api/review/approve_all", methods=["POST"])
def api_approve_all():
    """One control to rule them all: apply a single markup (+FX conversion) to
    every mismatch in the current scope and approve them together."""
    d = request.get_json(force=True) or {}
    markup = d.get("markup_pct")
    markup = 0.0 if markup in ("", None) else float(markup)
    ref = d.get("ref", "live")
    convert = d.get("convert", True)
    kind = d.get("kind", "mismatch")
    brands = [b for b in (d.get("brands") or []) if b]
    state = {"mismatch": "mismatch", "error": "error", "resolved": "matched"}.get(kind, "mismatch")
    where, params = "state=%s", [state]
    if brands:
        where += " AND brand IN (" + ",".join(["%s"] * len(brands)) + ")"
        params += brands
    con = store.connect()
    rows = con.execute(f"SELECT id, base_price, live_price, currency FROM products "
                       f"WHERE {where}", params).fetchall()
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    n = 0
    for r in rows:
        final = store_compute_final(r["base_price"], r["live_price"], r["currency"],
                                    ref, markup, None, convert)
        con.execute("UPDATE products SET decision='approved',markup_pct=%s,custom_price=NULL,"
                    "ref=%s,final_price=%s,decided_at=%s WHERE id=%s",
                    (markup, ref, final, now, r["id"]))
        n += 1
    con.commit()
    con.close()
    return jsonify(ok=True, approved=n, markup=markup, convert=bool(convert))


@app.route("/api/fx")
def api_fx():
    return jsonify(rates=fx.snapshot(("USD", "CAD", "INR")))


@app.route("/api/insights")
def api_insights():
    """Aggregates for the Power-BI-style home dashboard."""
    con = store.connect()
    c = store.counts()
    top_mis = con.execute(
        "SELECT brand, COUNT(*) c FROM products WHERE state='mismatch' AND brand<>'' "
        "GROUP BY brand ORDER BY c DESC LIMIT 10").fetchall()
    top_prod = con.execute(
        "SELECT brand, COUNT(*) c FROM products WHERE brand<>'' "
        "GROUP BY brand ORDER BY c DESC LIMIT 10").fetchall()
    agg = con.execute(
        "SELECT COALESCE(SUM(CASE WHEN delta>0 THEN delta ELSE 0 END),0) over_, "
        "COALESCE(SUM(CASE WHEN delta<0 THEN -delta ELSE 0 END),0) under_, "
        "COALESCE(AVG(ABS(delta)),0) avgd FROM products WHERE state='mismatch'").fetchone()
    vendors = con.execute(
        "SELECT COUNT(DISTINCT brand) c FROM products WHERE brand<>''").fetchone()["c"]
    approved_val = con.execute(
        "SELECT COALESCE(SUM(final_price),0) v FROM products WHERE decision='approved'"
        ).fetchone()["v"]
    con.close()
    return jsonify(
        counts=c, vendors=vendors,
        top_mismatch=[{"brand": r["brand"], "count": r["c"]} for r in top_mis],
        top_products=[{"brand": r["brand"], "count": r["c"]} for r in top_prod],
        exposure={"over": float(agg["over_"] or 0), "under": float(agg["under_"] or 0),
                  "avg": float(agg["avgd"] or 0)},
        approved_value=float(approved_val or 0),
        fx=fx.snapshot(("USD", "CAD", "INR")))


@app.route("/api/review/rerun", methods=["POST"])
def api_rerun():
    with RERUN_LOCK:
        if RERUN["running"]:
            return jsonify(ok=False, error="re-run already running"), 409
    d = request.get_json(silent=True) or {}
    con = store.connect()
    if d.get("rows"):
        ids = [int(x) for x in d["rows"]]
    else:
        ids = [r["id"] for r in con.execute("SELECT id FROM products WHERE state='error'").fetchall()]
    con.close()
    if not ids:
        return jsonify(ok=False, error="no error rows to re-run"), 400
    with RERUN_LOCK:
        RERUN.update(running=True, total=len(ids), done=0, recovered=0,
                     still_error=0, message=f"Slow re-run of {len(ids)} rows...")
    threading.Thread(target=_rerun_worker, args=(ids,), daemon=True).start()
    return jsonify(ok=True, total=len(ids))


def _rerun_worker(ids):
    fetcher = pt.Fetcher(timeout=15, cooldown_range=(4.0, 8.0), quiet=True)
    con = store.connect()
    recovered = still = 0
    try:
        for pid in ids:
            with RERUN_LOCK:
                if not RERUN["running"]:
                    break
            it = con.execute("SELECT url,platform,custom_regex,base_price,brand "
                             "FROM products WHERE id=%s", (pid,)).fetchone()
            if it is None:
                continue
            try:
                live, currency = pt.extract_row(fetcher, it["url"], it["platform"],
                                                it["custom_regex"] or None)
                if live is None:
                    raise ValueError("price not found")
            except Exception as exc:
                con.execute("UPDATE products SET rerun_status=%s,rerun_at=%s WHERE id=%s",
                            (f"Fetch Error: {exc}", time.strftime("%Y-%m-%d %H:%M:%S"), pid))
                con.commit()
                still += 1
                with RERUN_LOCK:
                    RERUN["done"] += 1
                    RERUN["still_error"] = still
                continue
            base = it["base_price"]
            cur = currency or "UNKNOWN"
            live_inr = fx.to_inr(live, cur)
            if base is not None and abs(live_inr - base) <= _match_tol(base, cur):
                state, st = "matched", f"Price Matched ({cur})"
            else:
                state, st = "mismatch", f"Price Mismatch! ({cur})"
            delta = (live_inr - base) if base is not None else None
            con.execute("UPDATE products SET state=%s,status=%s,live_price=%s,currency=%s,"
                        "delta=%s,rerun_status=%s,rerun_at=%s WHERE id=%s",
                        (state, st, live, cur, delta, st,
                         time.strftime("%Y-%m-%d %H:%M:%S"), pid))
            con.commit()
            recovered += 1
            with RERUN_LOCK:
                RERUN["done"] += 1
                RERUN["recovered"] = recovered
    finally:
        con.close()
        with RERUN_LOCK:
            RERUN["running"] = False
            RERUN["message"] = f"Done. {recovered} recovered, {still} still erroring."


@app.route("/api/review/rerun_status")
def api_rerun_status():
    with RERUN_LOCK:
        return jsonify(dict(RERUN))


@app.route("/api/review/rerun_stop", methods=["POST"])
def api_rerun_stop():
    with RERUN_LOCK:
        RERUN["running"] = False
    return jsonify(ok=True)


@app.route("/api/review/push", methods=["POST"])
def api_push():
    d = request.get_json(force=True) or {}
    pid = d.get("row")
    con = store.connect()
    it = con.execute("SELECT brand,url,final_price,decision FROM products WHERE id=%s",
                     (int(pid),)).fetchone()
    if it is None:
        con.close()
        return jsonify(ok=False, error="unknown row"), 404
    if it["decision"] != "approved":
        con.close()
        return jsonify(ok=False, error="approve a final price first"), 400
    res = push_price_to_shopify(it["brand"], it["url"], it["final_price"])
    con.execute("UPDATE products SET shopify_status=%s,shopify_at=%s WHERE id=%s",
                (res["status"], time.strftime("%Y-%m-%d %H:%M:%S"), int(pid)))
    con.commit()
    con.close()
    return jsonify(ok=res["ok"], status=res["status"])


@app.route("/api/review/push_all", methods=["POST"])
def api_push_all():
    with PUSH_LOCK:
        if PUSH["running"]:
            return jsonify(ok=False, error="a push is already running"), 409
    d = request.get_json(silent=True) or {}
    brand = (d.get("brand") or "").strip()
    con = store.connect()
    q = "SELECT id FROM products WHERE decision='approved'" + (" AND brand=%s" if brand else "")
    ids = [r["id"] for r in con.execute(q, (brand,) if brand else ()).fetchall()]
    con.close()
    if not ids:
        return jsonify(ok=False, error="no approved rows to push" + (f" for {brand}" if brand else "")), 400
    with PUSH_LOCK:
        PUSH.update(running=True, total=len(ids), done=0, ok=0, failed=0,
                    message=f"Pushing {len(ids)} approved price(s)...")
    threading.Thread(target=_push_all_worker, args=(ids,), daemon=True).start()
    return jsonify(ok=True, total=len(ids))


def _push_all_worker(ids):
    con = store.connect()
    ok = fail = 0
    try:
        for pid in ids:
            with PUSH_LOCK:
                if not PUSH["running"]:
                    break
            it = con.execute("SELECT brand,url,final_price FROM products WHERE id=%s", (pid,)).fetchone()
            if it is None:
                continue
            res = push_price_to_shopify(it["brand"], it["url"], it["final_price"])
            con.execute("UPDATE products SET shopify_status=%s,shopify_at=%s WHERE id=%s",
                        (res["status"], time.strftime("%Y-%m-%d %H:%M:%S"), pid))
            con.commit()
            ok, fail = (ok + 1, fail) if res["ok"] else (ok, fail + 1)
            with PUSH_LOCK:
                PUSH["done"] += 1
                PUSH["ok"], PUSH["failed"] = ok, fail
                PUSH["message"] = f"Pushing... {ok} ok, {fail} failed"
    finally:
        con.close()
        with PUSH_LOCK:
            PUSH["running"] = False
            PUSH["message"] = f"Done. {ok} pushed, {fail} failed."


@app.route("/api/review/push_status")
def api_push_status():
    with PUSH_LOCK:
        return jsonify(dict(PUSH))


# ---- Integrations API ----

@app.route("/api/integrations")
def api_integrations():
    con = store.connect()
    brands = con.execute(
        "SELECT brand, COUNT(*) c, COUNT(*) FILTER (WHERE state='mismatch') m "
        "FROM products WHERE brand<>'' GROUP BY brand ORDER BY c DESC").fetchall()
    cfgs = {r["brand"]: dict(r) for r in con.execute("SELECT * FROM integrations").fetchall()}
    con.close()
    out = []
    for b in brands:
        c = cfgs.get(b["brand"], {})
        out.append({"brand": b["brand"], "products": b["c"], "mismatches": b["m"] or 0,
                    "shop_domain": c.get("shop_domain", ""),
                    "api_version": c.get("api_version", "2024-10"),
                    "dry_run": bool(c.get("dry_run", 0)),
                    "has_token": bool(c.get("access_token"))})
    return jsonify(brands=out)


@app.route("/api/integrations/save", methods=["POST"])
def api_int_save():
    d = request.get_json(force=True) or {}
    brand = (d.get("brand") or "").strip()
    if not brand:
        return jsonify(ok=False, error="missing brand"), 400
    con = store.connect()
    existing = con.execute("SELECT access_token FROM integrations WHERE brand=%s", (brand,)).fetchone()
    token = (d.get("access_token") or "").strip()
    if not token and existing:  # keep existing token if field left blank
        token = existing["access_token"]
    con.execute("""INSERT INTO integrations(brand,shop_domain,access_token,api_version,dry_run,updated_at)
                   VALUES(%s,%s,%s,%s,%s,%s)
                   ON CONFLICT(brand) DO UPDATE SET shop_domain=excluded.shop_domain,
                     access_token=excluded.access_token, api_version=excluded.api_version,
                     dry_run=excluded.dry_run, updated_at=excluded.updated_at""",
                (brand, (d.get("shop_domain") or "").strip(), token,
                 (d.get("api_version") or "2024-10").strip(),
                 1 if d.get("dry_run") else 0, time.strftime("%Y-%m-%d %H:%M:%S")))
    con.commit()
    con.close()
    return jsonify(ok=True)


@app.route("/api/integrations/verify", methods=["POST"])
def api_int_verify():
    brand = ((request.get_json(silent=True) or {}).get("brand") or "").strip()
    if not brand:
        return jsonify(ok=False, status="missing brand")
    return jsonify(**verify_shopify(brand))


@app.route("/api/integrations/delete", methods=["POST"])
def api_int_delete():
    brand = ((request.get_json(silent=True) or {}).get("brand") or "").strip()
    con = store.connect()
    con.execute("DELETE FROM integrations WHERE brand=%s", (brand,))
    con.commit()
    con.close()
    return jsonify(ok=True)


# ---- Alerts API (price drops / spikes vs the previous run) ----

@app.route("/api/alerts")
def api_alerts():
    """Compare each product's latest price snapshot to its previous one and
    flag movements beyond `threshold` percent. direction = all|drop|spike."""
    try:
        threshold = abs(float(request.args.get("threshold", 5)))
    except (TypeError, ValueError):
        threshold = 5.0
    direction = request.args.get("direction", "all")
    brand = request.args.get("brand", "").strip()

    con = store.connect()
    params = []                       # placeholders in query order: [brand?], threshold
    bclause = ""
    if brand:
        bclause = "AND brand=%s"
        params.append(brand)
    params.append(threshold)
    rows = con.execute(f"""
        SELECT key, url, brand, live_price, prev, created_at, status,
               (live_price - prev) AS abs_change,
               ROUND(((live_price - prev) / prev * 100)::numeric, 2) AS pct
        FROM (
          SELECT key, url, brand, live_price, status, created_at,
            LAG(live_price) OVER (PARTITION BY key ORDER BY created_at) AS prev,
            ROW_NUMBER() OVER (PARTITION BY key ORDER BY created_at DESC) AS rn
          FROM price_history
          WHERE live_price IS NOT NULL
        ) t
        WHERE rn = 1 AND prev IS NOT NULL AND prev <> 0 {bclause}
          AND ABS((live_price - prev) / prev * 100) >= %s
        ORDER BY ABS((live_price - prev) / prev * 100) DESC
        LIMIT 1000
    """, params).fetchall()
    con.close()

    items = []
    for r in rows:
        pct = float(r["pct"]) if r["pct"] is not None else 0.0
        dir_ = "spike" if pct > 0 else "drop"
        if direction in ("drop", "spike") and dir_ != direction:
            continue
        d = dict(r)
        d["pct"] = pct
        d["abs_change"] = float(r["abs_change"]) if r["abs_change"] is not None else None
        d["direction"] = dir_
        d["created_at"] = str(d["created_at"])
        items.append(d)
    spikes = sum(1 for i in items if i["direction"] == "spike")
    drops = sum(1 for i in items if i["direction"] == "drop")
    return jsonify(items=items, total=len(items), spikes=spikes, drops=drops,
                   threshold=threshold)


@app.route("/api/alerts/brands")
def api_alerts_brands():
    con = store.connect()
    rows = con.execute(
        "SELECT brand, COUNT(DISTINCT key) c FROM price_history "
        "WHERE brand<>'' GROUP BY brand ORDER BY c DESC").fetchall()
    con.close()
    return jsonify(brands=[{"brand": r["brand"], "count": r["c"]} for r in rows])


# ---------------------------------------------------------------------------

def serve():
    ok, msg = store.ping()
    if not ok:
        print("[PriceSync] Supabase connection FAILED:", msg)
        print("[PriceSync] Set SUPABASE_DB_URL in .env (see .env.example).")
        raise SystemExit(1)
    store.init()  # ensure tables exist; the app only works on sheets you import
    print("[PriceSync] Supabase connected.")
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8080"))
    try:
        from waitress import serve as wserve
        print(f"[PriceSync] http://{host}:{port}")
        wserve(app, host=host, port=port, threads=int(os.environ.get("THREADS", "16")))
    except ImportError:
        app.run(host=host, port=port, threaded=True)


if __name__ == "__main__":
    serve()
